import logging
logging.basicConfig(format='[%(levelname) 5s/%(asctime)s] %(name)s: %(message)s',
                    level=logging.WARNING)

import os
from time import sleep
from telethon import TelegramClient, events, functions, types
import openpyxl
from openpyxl import styles
import configparser
from datetime import datetime, timedelta

# Settings and global vars

config = configparser.ConfigParser()
config.read('config.ini')

api_id = config.getint('BOT', 'api_id')
api_hash = config['BOT']['api_hash']
user_bot_phone = config['BOT']['user_bot_phone']

bot_owner_id = config.getint('INVITES', 'bot_owner_id')
excel_file = config['INVITES']['excel_file']
aim_channel_id = config.getint('INVITES', 'aim_channel_id')
sleep_time = config.getint('INVITES', 'sleep_time')
max_per_day = config.getint('INVITES', 'max_per_day')
iteration_limit = 50 # Needs to avoid ban on importing contacts

limit_today = max_per_day
last_date = datetime.today().date() -  timedelta(days=1)

message = ''

try:
    file_exists = os.path.exists(excel_file)
except Exception as e:
    print(e)
    file_exists = False

try:
    message_file = open('message.txt', 'r')
    message = message_file.read()
    message_file.close()
except Exception as e:
    print(e)
    message = 'Hello!'

# Functions

def getContactsFromExcel(filename, count):
    excel = openpyxl.load_workbook(filename)
    excel_numbers = excel.active
    stop_cell = excel_numbers['B1']
    phone_numbers = []

    if not isinstance(stop_cell.value, int):
        stop_cell.value = 0
        excel.save(filename)

    last_stop = stop_cell.value

    row = last_stop
    for row in range(last_stop + 1, excel_numbers.max_row + 1):
        cell = excel_numbers.cell(row=row, column=1)
        
        if cell.value is not None:
            phone_numbers.append(cell.value)
            cell.font = styles.Font(name='Calibri', 
                                    charset=204, 
                                    family=2.0, 
                                    b=True, 
                                    color='74C43F', 
                                    sz=11.0)

        if len(phone_numbers) >= count:
            break

    stop_cell.value = row
    excel.save(filename)

    users_contacts = list((lambda i: types.InputPhoneContact(
             client_id = 0,
             phone = phone_numbers[i],
             first_name = 'Invited',
             last_name = 'User'
            ))(i) for i in range(len(phone_numbers)))

    return users_contacts

async def getUsers(contacts):
    users_to_invite = await bot(functions.contacts.ImportContactsRequest(contacts))

    users_entities = []
    for i in range(len(users_to_invite.imported)):
        entity = await bot.get_entity(types.PeerUser(users_to_invite.imported[i].user_id))
        users_entities.append(entity)
    
    return users_entities

async def deleteContacts(users):
    users_ids = [user.id for user in users]
    try:
        await bot(functions.contacts.DeleteContactsRequest(users_ids))
    except Exception as e:
        print(e)

async def inviteUsersToChannel(channel, participants, users):
    invited = 0

    try:
        channel = await bot.get_entity(types.PeerChannel(-1001221685276))
    except Exception as e:
        print(e)

    for user in users:
        try:
            await bot(functions.channels.EditAdminRequest(
                channel=channel,
                user_id=user.id,
                admin_rights=types.ChatAdminRights(other=True),
                rank=''
            ))

            rights = types.ChatAdminRights(False, False, False, False, False, False, False, False, False, False, False)

            await bot(functions.channels.EditAdminRequest(
                channel=channel,
                user_id=user.id,
                admin_rights=rights,
                rank=''
            ))

            invited += 1
            sleep(sleep_time)
        except Exception as e:
            print(e)

    return invited

async def sendMessages(users, message):
    sent = 0

    for user in users:
        try:
            await bot.send_message(user, message)
            sent += 1
        except Exception as e:
            print(e)
        
        sleep(sleep_time)

    return sent


# Handlers

bot = TelegramClient(config['BOT']['bot_session'], api_id, api_hash).start(user_bot_phone)

@bot.on(events.NewMessage(incoming=True, from_users=[bot_owner_id], pattern='/excel'))
async def excel_event_handler(event):
    global excel_file, file_exists

    file = event.file

    if file and file.name.endswith('.xlsx'):
        if (file_exists):
            await event.respond('Старый список (использованные номера помечены зеленым).', 
                                file=excel_file)

        try:
            await event.download_media(file=excel_file)
        except Exception as e:
            print(e)
            await event.respond('Ошибка при загрузке файла')
            return

        await event.respond('Список номеров обновлен!')
    else:
        await event.respond('Для загрузки нового списка номеров, '
                            'загрузите файл .xlsx и отправьте с командой /excel')

@bot.on(events.NewMessage(incoming=True, from_users=[bot_owner_id], pattern='/invite'))
async def invite_event_handler(event):
    global excel_file, file_exists, limit_today, max_per_day, iteration_limit, last_date, aim_channel_id

    # If today is a new day, setting a new limit
    if last_date == datetime.today().date() - timedelta(days=1):
        limit_today = max_per_day
        last_date = datetime.today().date()
        
    if limit_today < 1:
        await event.respond('Лимит приглашений на сегодня исчерпан.')
        return

    if not file_exists:
        await event.respond('Загрузите список номеров')
        return

    try:
        aim_channel = await bot.get_entity(types.PeerChannel(aim_channel_id))
        
        # Getting channel participants
        offset = 0
        limit = 1
        participants_ids = []

        while True:
            participants = await bot(functions.channels.GetParticipantsRequest(
                aim_channel, types.ChannelParticipantsSearch(''), offset, limit, hash=0
            ))
            if not participants.users:
                break
            ids = [user.id for user in participants.users]
            participants_ids.extend(ids)
            offset += len(participants.users)

        total_invited = 0
        total_users_count = 0

        while limit_today > 0:
            users_contacts = getContactsFromExcel(excel_file, min(limit_today, iteration_limit))
            if len(users_contacts) < 1:
                await event.respond('Все номера из файла excel были использованы.')
                break

            users_entities = await getUsers(users_contacts)
            users_to_invite = []

            for user in users_entities:
                if user.id in participants_ids:
                    print(f'Уже в группе: {user.id}')
                else:
                    users_to_invite.append(user)


            users_count = len(users_to_invite)
            invited = await inviteUsersToChannel(aim_channel, participants, users_to_invite)
            limit_today -= invited   

            total_invited += invited
            total_users_count += users_count  

            await deleteContacts(users_entities)

        await event.respond(f'Было приглашено {total_invited} пользователей из '
                            f'{total_users_count} ({total_users_count - total_invited} не удалось).')
    except Exception as e:
        print(e)
        await event.respond('Произошла ошибка. Попробуйте через 30 с.')

@bot.on(events.NewMessage(incoming=True, from_users=[bot_owner_id], pattern='/sendmessages'))
async def sendmessages_event_handler(event):
    global excel_file, file_exists, limit_today, max_per_day, iteration_limit, last_date, aim_channel_id, message

    # If today is a new day, setting a new limit
    if last_date == datetime.today().date() - timedelta(days=1):
        limit_today = max_per_day
        last_date = datetime.today().date()

    if limit_today < 1:
        await event.respond('Лимит приглашений на сегодня исчерпан.')
        return

    if not file_exists:
        await event.respond('Загрузите список номеров')
        return

    try:
        total_sent = 0
        total_users_count = 0

        while limit_today > 0:
            users_contacts = getContactsFromExcel(excel_file, min(limit_today, iteration_limit))
            if len(users_contacts) < 1:
                await event.respond('Все номера из файла excel были использованы.')
                break

            users_to_invite = await getUsers(users_contacts)

            users_count = len(users_to_invite)
            sent = await sendMessages(users_to_invite, message)
            limit_today -= sent
            
            total_sent += sent
            total_users_count += users_count

            await deleteContacts(users_to_invite)

        await event.respond(f'Было разослано {total_sent} сообщений из '
                            f'{total_users_count} ({total_users_count - total_sent} не удалось).')
    except Exception as e:
        print(e)
        await event.respond('Произошлка ошибка')

@bot.on(events.NewMessage(incoming=True, from_users=[bot_owner_id], pattern=r'\/message'))
async def message_event_handler(event):
    global message

    if event.raw_text == '/message':
        await event.respond('Необходим текст для рассылки после /message')
        return

    try:
        new_message = event.raw_text[8:].strip()

        message_file = open('message.txt', 'w', encoding='utf-8')
        message_file.write(new_message)
        message_file.close()

        message = new_message

        await event.respond(f'Сообщение обновлено!\n{message}')
    except Exception as e:
        print(e)
        await event.respond('Произошла ошибка при обновлении сообщения')

@bot.on(events.NewMessage(incoming=True, from_users=[bot_owner_id], pattern='/help'))
async def help_event_handler(event):
    await event.respond('/excel (+файл .xlsx) - обновление списка номеров\n'
                        '/invite - пригласить людей по номерам в канал\n'
                        '/message (+сообщение) - новое сообщение для рассылки\n'
                        '/sendmessages - начать рассылку\n'
                        '/limit - узнать оставшийся на сегодня лимит приглашений\n'
                        '/sleep - узнать задержку между приглашениями\n'
                        '/channel - узнать название и ID канала\n'
                        '**\nСледующие команды стоит использовать с осторожностью:**\n'
                        '/setlimit (+число) - установить новый лимит\n'
                        '/setsleep (+число) - установить новую задержку\n'
                        '/setchannel (+ID канала) - установить новый канал\n'
                        'Перешлите сообщение из канала боту, чтобы узнать ID этого канала')

@bot.on(events.NewMessage(incoming=True, from_users=[bot_owner_id], pattern='/limit'))
async def limit_event_handler(event):
    global limit_today
    await event.respond(f'На сегодня доступно {limit_today} приглашений.')

@bot.on(events.NewMessage(incoming=True, from_users=[bot_owner_id], pattern='/setlimit'))
async def setlimit_event_handler(event):
    global limit_today, max_per_day, config

    try:
        new_limit = int(event.raw_text.lstrip('/setlimit').strip())

        config['INVITES']['max_per_day'] = str(new_limit)
        with open('config.ini', 'w') as configfile:
            config.write(configfile)
            
        limit_today += new_limit - max_per_day
        max_per_day = new_limit
    except ValueError as value_error:
        print(value_error)
        await event.respond('Для обновления ежедневного лимита нужно ввести целое число.')
        return
    except Exception as e:
        print(e)
        await event.respond('Произошла ошибка при обновлении лимита. Попробуйте снова.')
        return

    await event.respond(f'Лимит обновлен! Теперь {max_per_day}')

@bot.on(events.NewMessage(incoming=True, from_users=[bot_owner_id], pattern='/sleep'))
async def sleep_event_handler(event):
    global sleep_time
    await event.respond(f'Текущая задержка между приглашениями: {sleep_time} с.')

@bot.on(events.NewMessage(incoming=True, from_users=[bot_owner_id], pattern='/setsleep'))
async def setsleep_event_handler(event):
    global sleep_time, config

    try:
        new_sleep_time = int(event.raw_text.lstrip('/setsleep').strip())

        if new_sleep_time < 0:
            await event.respond('Нужно ввести 0 или положительное число.')
            return

        config['INVITES']['sleep_time'] = str(new_sleep_time)
        with open('config.ini', 'w') as configfile:
            config.write(configfile)
            
        sleep_time = new_sleep_time
    except ValueError as value_error:
        print(value_error)
        await event.respond('Для обновления задержки нужно ввести целое число.')
        return
    except Exception as e:
        print(e)
        await event.respond('Произошла ошибка при обновлении задержки. Попробуйте снова.')
        return

    await event.respond(f'Задержка обновлена! Теперь {sleep_time}')

@bot.on(events.NewMessage(incoming=True, from_users=[bot_owner_id], pattern='/channel'))
async def channel_event_handler(event):
    global aim_channel_id
    print(aim_channel_id)
    channel = await bot.get_entity(types.PeerChannel(aim_channel_id))

    await event.respond(f'Канал: {channel.title} ({aim_channel_id})')

@bot.on(events.NewMessage(incoming=True, from_users=[bot_owner_id], pattern='/setchannel'))
async def setchannel_event_handler(event):
    global aim_channel_id, config

    try:
        channel_str = event.raw_text.lstrip('/setchannel').strip()
        if not str(channel_str.startswith('-100')):
            await event.respond('ID канала должен начинаться с "-100"')
            return

        new_channel_id = int(channel_str)

        # Check if channel exists
        channel = await bot.get_entity(types.PeerChannel(new_channel_id))

        config['INVITES']['aim_channel_id'] = str(new_channel_id)
        with open('config.ini', 'w') as configfile:
            config.write(configfile)

        aim_channel_id = new_channel_id
    except ValueError as value_error:
        print(value_error)
        await event.respond('Канал с таким ID не найден.')
        return
    except Exception as e:
        print(e)
        await event.respond('Произошла ошибка при обновлении канала.\nПопробуйте снова.')
        return

    await event.respond(f'Новый канал: {channel.title}')

@bot.on(events.NewMessage(incoming=True, from_users=[bot_owner_id], forwards=True))
async def msgforward_event_handler(event):
    await event.respond(f'ID канала: -100{event.fwd_from.from_id.channel_id}')


bot.start()
bot.run_until_disconnected()